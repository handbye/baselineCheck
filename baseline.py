#coding: utf-8

# 基线自动化脚本,目前只能对linux进行基线检查

import nmap
from openpyxl import load_workbook
from fabric.api import *
from fabric.context_managers import *
from fabric.contrib.console import confirm
import paramiko
import time
import argparse


# 从excel中读取ip,用户名,密码
def readip(filename):
    iplist = []
    portlist = []
    rootpasswddict = []
    usernamedict = []
    passwddict = []
    wb = load_workbook(filename)
    sheet_names = wb.sheetnames
    sheet_name = sheet_names[0]
    ws = wb[sheet_name]
    rows = ws.max_row
    columns = ws.max_column
    for r in range(2,rows+1):
        ip = ws.cell(r, 1).value
        port = ws.cell(r,2).value
        rootpasswd = ws.cell(r,3).value
        username = ws.cell(r,4).value
        passwd = ws.cell(r,5).value
        iplist.append(ip)
        portlist.append(port)
        rootpasswddict.append(rootpasswd)
        usernamedict.append(username)
        passwddict.append(passwd)
    return iplist,portlist,rootpasswddict,usernamedict,passwddict


# 借助nmap判断系统类型
def judgesystem(ip,port):
    nm = nmap.PortScanner()
    nm.scan(ip,str(port),arguments='-Pn')
    try:
        state = nm[ip]['tcp'][port]['state']
        protocol = nm[ip]['tcp'][port]['name']
        if state == 'open' and protocol == 'ssh':
            return 0
        elif state == 'open' and protocol == 'rdp':
            return 1
        else:
            return 2
    except KeyError as exkey:
        print("[!]the host: %s may not up."%ip)


# fabric需要全局定义hostname等参数，所以使用了setenv函数实现
def setenv(hostname,username,password,port):
    env.host_string = hostname
    env.port = str(port)
    env.user = username
    env.password = password
    return env


# linux上传基线脚本并执行
def linux_put(hostname,username,password,port,lfile):
    env = setenv(hostname,username,password,port)
    with cd("/tmp"):
        with settings(warn_only=True):
            result = put(lfile,"/tmp")
        if result.failed and not confirm("put file failed, Continue[Y/N]?"):
            abort("Aborint file put task!")


def linux_get(hostname,username,password,port,r_result,l_result):
    env = setenv(hostname,username,password,port)
    with cd("/tmp"):
        with settings(warn_only=True):
            result = get(r_result,l_result)
        if result.failed and not confirm("put file failed, Continue[Y/N]?"):
            abort("Aborint file get task!")


# 执行linux基线检查脚本
def linux_check(hostname,username,userpass,port,rootpass,command):
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(hostname,port, username,userpass)
    shell = client.invoke_shell()
    shell.send("su\n")
    time.sleep(1)
    shell.send("{0}\n".format(rootpass))
    time.sleep(2)
    shell.send("cd /tmp/Linux\n")
    shell.send("bash {0} {1}\n".format(command,hostname))
    time.sleep(3)  #时间必须足够长，否则会造成命令没执行完毕就退出的情况
    receive_buffer = shell.recv(102400).decode('utf-8')
    print(receive_buffer)
    shell.send("cd /tmp\n")
    shell.send("rm -rf Linux/\n")
    client.close()


def arg():
    parse = argparse.ArgumentParser(description="a baseline check linux_baseline_check")
    parse.add_argument("-x", "--xlsx", help="the target excel path and name", required=True)
    parse.add_argument("-l","--lfile",help="loacl file or local directory",required=True)
    parse.add_argument("-c","--command",help='command for run check linux_baseline_check, default is "check_server_linux.sh"',
                       default='check_server_linux.sh')
    parse.add_argument("-ld","--ldirectory",help='save check result to local directory, default is "./"',
                       default='./')
    args = parse.parse_args()
    return args


if __name__ == '__main__':
    argss = arg()
    ips = readip(argss.xlsx)[0]
    ports = readip(argss.xlsx)[1]
    rootpasswors = readip(argss.xlsx)[2]
    usernames = readip(argss.xlsx)[3]
    userpasswors = readip(argss.xlsx)[4]
    for ip in ips:
        port = ports[ips.index(ip)]
        l = judgesystem(ip,port)
        if l == 0:
            print('-' * 80)
            print('check the {0} '.format(ip))
            print("*" * 80)
            username = usernames[ips.index(ip)]
            userpass = userpasswors[ips.index(ip)]
            setenv(ip,username,userpass,port)
            linux_put(ip,username,userpass,port,argss.lfile)
            rootpass = rootpasswors[ips.index(ip)]
            linux_check(ip,username,userpass,port,rootpass,argss.command)
            print("-" * 80)
            print("check {0} finish!".format(ip))
            print("*" * 80)
            r_result = ip+"_linux_chk.xml"
            print("down check result to local")
            linux_get(ip,username,userpass,port,r_result,argss.ldirectory)
            print("-" * 80)